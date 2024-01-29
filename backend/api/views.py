import json
from .models import Instiute, Student
from django.http import HttpResponse
from django.core import serializers
from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt,csrf_protect
from django.contrib import messages
from django.shortcuts import render
from openpyxl import load_workbook

# Create your views here.

def paymentlogin(request):
    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        date_of_birth = request.POST.get('date_of_birth')

        student = Student.objects.filter(student_id=student_id).first()

        if student and student.date_of_birth == date_of_birth:
            authenticated_student  = authenticate(request, student_id=student_id, date_of_birth=date_of_birth)
            if authenticated_student:
                login(request, authenticated_student)
                return HttpResponse(json.dumps({"msg": " your details updated successfully."}),content_type="application/json",)
            else:
                messages.error(request, "invalid")
        else:
            messages.error(request, "invalid")

@csrf_exempt
def add_student(request):
    if request.method == "POST":
        student = Student()
        student.name = request.POST.get("name")
        institute_name = request.POST.get("institute") 
        instiute = Instiute.objects.get(institute_name=institute_name)
        student.institute = instiute
        student.course = request.POST.get("course")
        student.email = request.POST.get("email")
        student.division = request.POST.get("division")
        student.father_name = request.POST.get("father_name")
        student.phone_number = request.POST.get("phone_number")
        student.parent_number = request.POST.get("parent_number")
        student.date_of_birth = request.POST.get("date_of_birth")
        student.gender = request.POST.get("gender")
        student.fees_paid = request.POST.get("fees_paid")
        student.fees_unpaid = request.POST.get("fees_unpaid")
        student.address = request.POST.get("address")
        student.city = request.POST.get("city")
        student.state = request.POST.get("state")
        student.pincode = request.POST.get("pincode")
        student.save()
        return HttpResponse(json.dumps({"msg": " your details updated successfully."}),content_type="application/json",)
    else:
        return HttpResponse(json.dumps({"msg": " there has been an error"}),content_type="application/json",)
    
@csrf_exempt
def findstudents(request):
    if request.method == "GET":
        institute = request.GET.get("institute")
        course = request.GET.get("course")
        division = request.GET.get("division")
        sem = request.GET.get("sem")
        allstudents = Student.objects.filter(institute=institute, course=course, division=division, study_sem=sem)
        data = serializers.serialize("json" ,allstudents)
        return HttpResponse(data, content_type="application/json")


def import_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            name, institute, course, email, division, father_name, phone_number, parent_number, date_of_birth, gender, fees_paid, fees_unpaid, address, city, state, pincode = row
            Student.objects.create(name=name, institute=institute, course=course, email=email, date_of_birth=date_of_birth, gender=gender, fees_paid=fees_paid, fees_unpaid=fees_unpaid, address=address, city=city, state=state, pincode=pincode, division=division, father_name=father_name, phone_number=phone_number, parent_number=parent_number)

        return HttpResponse(json.dumps({"msg": " your details updated successfully."}),content_type="application/json",)

    return HttpResponse(json.dumps({"msg": " your details updated successfully."}),content_type="application/json",)